# Create your views here.
from django.db.models import Q
from rest_framework.decorators import api_view, authentication_classes
from datetime import datetime, timedelta
from myapp.auth.authentication import AdminTokenAuthtication
from myapp.handler import APIResponse
from myapp.models import Airline  # 新增导入
from myapp.permission.permission import isDemoAdminUser
from myapp.serializers import AirlineSerializer  # 假设你已经创建了对应的序列化器




@api_view(['GET'])
def list_api(request):
    if request.method == 'GET':
        keyword = request.GET.get("keyword", '')
        airlines = Airline.objects.filter(id__contains=keyword)  # 修改为查询Airline
        ids_arr=[]
        serializer = AirlineSerializer(airlines, many=True)  # 使用AirlineSerializer
        for i in serializer.data:
            end_time=datetime.strptime(i['end_time'],"%Y-%m-%d %H:%M")
            time_diff = datetime.now() - end_time.replace(tzinfo=None)
            if time_diff > timedelta(hours=12):
                ids_arr += [i['id']]
        Airline.objects.filter(Q(id__in=ids_arr)).delete()
        return APIResponse(code=0, msg='查询成功', data=serializer.data)

@api_view(['POST'])
@authentication_classes([AdminTokenAuthtication])
def create(request):
    if isDemoAdminUser(request):
        return APIResponse(code=1, msg='演示帐号无法操作')

    serializer = AirlineSerializer(data=request.data)  # 使用AirlineSerializer
    print(serializer)
    if serializer.is_valid():
        serializer.save()
        return APIResponse(code=0, msg='创建成功', data=serializer.data)

    return APIResponse(code=1, msg='创建失败')

@api_view(['POST'])
@authentication_classes([AdminTokenAuthtication])
def update(request):
    if isDemoAdminUser(request):
        return APIResponse(code=1, msg='演示帐号无法操作')

    try:
        pk = request.GET.get('id', -1)
        airline = Airline.objects.get(pk=pk)  # 修改为查询Airline
    except Airline.DoesNotExist:
        return APIResponse(code=1, msg='对象不存在')

    serializer = AirlineSerializer(airline, data=request.data)  # 使用AirlineSerializer
    if serializer.is_valid():
        serializer.save()
        return APIResponse(code=0, msg='更新成功', data=serializer.data)

    return APIResponse(code=1, msg='更新失败'+str(serializer.errors))

@api_view(['POST'])
@authentication_classes([AdminTokenAuthtication])
def delete(request):
    print(request)
    if isDemoAdminUser(request):
        return APIResponse(code=1, msg='演示帐号无法操作')

    try:
        ids = request.GET.get('ids')

        ids_arr = ids.split(',')
        Airline.objects.filter(Q(id__in=ids_arr)).delete()  # 修改为删除Airline
    except Airline.DoesNotExist:
        return APIResponse(code=1, msg='对象不存在')
    except Exception as e:
        print(e)
    return APIResponse(code=0, msg='删除成功')


@api_view(['POST'])
@authentication_classes([AdminTokenAuthtication])
def import_excel(request):
    if isDemoAdminUser(request):
        return APIResponse(code=1, msg='演示帐号无法操作')
    try:
        file = request.FILES['file']
        if not file:
            return APIResponse(code=1, msg='请上传文件')

        # 新增 Excel 处理逻辑
        from openpyxl import load_workbook
        import json

        wb = load_workbook(file)
        ws = wb.active

        success = 0
        fail = 0
        errors = []

        # 跳过标题行（第一行）
        for row in ws.iter_rows(min_row=2):
            try:
                # 解析数据
                people_list = row[5].value.split(',')  # 拆分逗号分隔的字符串

                airline_data = {
                    'id': row[0].value,
                    'begin_time': str(row[1].value),
                    'end_time': str(row[2].value),
                    'begin_ariport': row[3].value,
                    'end_ariport': row[4].value,
                    'people': json.dumps(people_list)  # 转换为 JSON 字符串
                }

                # 使用序列化器验证并保存
                serializer = AirlineSerializer(data=airline_data)
                if serializer.is_valid():
                    serializer.save()
                    success += 1
                else:
                    raise ValueError(serializer.errors)

            except Exception as e:
                fail += 1
                errors.append(f"第{row[0].row}行错误: {str(e)}")

        return APIResponse(code=0, msg=f'导入完成（成功{success}条，失败{fail}条）', data={
            'success': success,
            'fail': fail,
            'errors': errors
        })

    except KeyError:
        return APIResponse(code=1, msg='请上传文件')
    except Exception as e:
        return APIResponse(code=1, msg=f'导入失败: {str(e)}')
